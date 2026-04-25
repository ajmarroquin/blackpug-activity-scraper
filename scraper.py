import time
import openpyxl
import os
import re
import shutil
import string
import subprocess
import sys
import traceback

from datetime import datetime

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from collections import defaultdict

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException


def build_driver():
    """Build a Chrome webdriver using Selenium Manager (cross-platform).

    Selenium >= 4.11 ships with Selenium Manager, which auto-discovers an
    installed Chrome/Chromium and downloads a matching chromedriver into
    ~/.cache/selenium. No manual driver setup required on Linux, macOS, or
    Windows.

    Optional environment variables:
      CHROME_BINARY  -- absolute path to a specific Chrome/Chromium binary
                        (e.g. a Chrome for Testing build).
      HEADLESS=1     -- run in headless mode (note: you log in interactively,
                        so the default is headed).
    """
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")

    chrome_binary = os.environ.get("CHROME_BINARY")
    if chrome_binary:
        chrome_binary = os.path.expanduser(chrome_binary)
        options.binary_location = chrome_binary
    else:
        chrome_binary = _autodetect_chrome_binary()
        if chrome_binary:
            options.binary_location = chrome_binary

    _abort_if_snap_chromium(chrome_binary)

    if os.environ.get("HEADLESS") in ("1", "true", "True"):
        options.add_argument("--headless=new")

    return webdriver.Chrome(options=options)


def _autodetect_chrome_binary():
    """Find a non-snap Chrome/Chromium binary on PATH.

    Selenium Manager will otherwise happily pick up /snap/bin/chromium on
    Ubuntu, which is sandboxed and silently fails to navigate (browser opens
    to ``data:,`` and never loads the URL).
    """
    candidates = [
        "google-chrome",
        "google-chrome-stable",
        "chromium",
        "chromium-browser",
        "chrome",
    ]
    for name in candidates:
        path = shutil.which(name)
        if path and "/snap/" not in os.path.realpath(path):
            return path
    return None


def _abort_if_snap_chromium(binary_path):
    """Hard-fail with a clear message if we're about to launch snap Chromium."""
    if not binary_path:
        # Let Selenium Manager try; if it picks snap chromium we can't tell here.
        # Check the default snap path explicitly.
        snap_path = "/snap/bin/chromium"
        if os.path.exists(snap_path) and not shutil.which("google-chrome") \
                and not shutil.which("google-chrome-stable"):
            _print_snap_error()
            sys.exit(1)
        return

    real = os.path.realpath(binary_path)
    if "/snap/" in real:
        _print_snap_error(real)
        sys.exit(1)


def _print_snap_error(path=""):
    print("=" * 70)
    print("❌ Snap-confined Chromium detected" + (f" at {path}" if path else ""))
    print("=" * 70)
    print("Snap Chromium is sandboxed and cannot be driven by Selenium —")
    print("the browser opens to 'data:,' and never navigates.")
    print()
    print("Fix (recommended): install the official Google Chrome .deb:")
    print("  wget https://dl.google.com/linux/direct/google-chrome-stable_current_amd64.deb")
    print("  sudo apt install ./google-chrome-stable_current_amd64.deb")
    print()
    print("Or download Chrome for Testing and set CHROME_BINARY:")
    print("  https://googlechromelabs.github.io/chrome-for-testing/")
    print("  export CHROME_BINARY=/path/to/chrome")
    print("=" * 70)

def auto_fit_columns(ws):
    for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

def extract_event_rows(driver, filter_option="current_year"):
    wait = WebDriverWait(driver, 10)
    rows = driver.find_elements(By.CSS_SELECTOR, "div[onclick*='toggle']")
    print(f"📦 Found {len(rows)} registration rows...")

    event_data = defaultdict(lambda: {"booked": [], "not_booked": []})

    for row in rows:
        try:
            event_title = row.text.strip()
            
            # Apply filtering based on user choice
            if filter_option == "future" and not is_future_event(event_title):
                continue  # ⏩ Skip if not a future event
            elif filter_option == "current_year" and not is_current_year_event(event_title):
                continue  # ⏩ Skip if not from current year
            elif filter_option == "last_year" and not is_last_year_event(event_title):
                continue  # ⏩ Skip if not from last year
            driver.execute_script("arguments[0].click();", row)
            time.sleep(0.4)

            container = row.find_element(By.XPATH, "following-sibling::div[1]")
            fields = container.find_elements(By.CSS_SELECTOR, "div.col-xs-12.col-sm-6")

            record = {}
            is_not_booked = "Not Booked" in event_title
            
            # Extract registration number from event title (the number in parentheses)
            reg_number_match = re.search(r'\((\d+)\)$', event_title)
            registration_number = reg_number_match.group(1) if reg_number_match else ""
            
            for field in fields:
                try:
                    label = field.find_element(By.CLASS_NAME, "col-xs-4").text.strip()
                    value = field.find_element(By.CLASS_NAME, "col-xs-8").text.strip()
                    
                    # Skip Balance Due column
                    if label == "Balance Due":
                        continue
                        
                    # Ensure label is always a string
                    label = str(label).strip()
                    if not label:
                        continue  # Skip empty labels
                        
                    # Convert Participants to number
                    if label == "Participants":
                        try:
                            value = int(value)
                        except ValueError:
                            value = 0
                    
                    record[label] = value
                except:
                    continue
            
            # Add registration number as a separate field
            if registration_number:
                record["Registration Number"] = int(registration_number)  # Store as number, not string

            # Extract clean event name (remove date prefix and registration number)
            event_base = event_title.split(":")[1].strip() if ":" in event_title else event_title
            # Remove "Not Booked" prefix if present
            if event_base.startswith("Not Booked "):
                event_base = event_base[11:]  # Remove "Not Booked " prefix
            # Remove registration number in parentheses
            event_key = re.sub(r"\s*\(\d+\)$", "", event_base)
            
            # Separate booked vs not booked
            if is_not_booked:
                event_data[event_key]["not_booked"].append(record)
            else:
                event_data[event_key]["booked"].append(record)

        except Exception as e:
            print("⚠️ Skipped a row due to error:", e)

    # Clean up duplicate registrations: remove people from not_booked if they're in booked
    debug_log = []
    debug_log.append(f"=== DEDUPLICATION DEBUG LOG - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
    
    for event_key, data in event_data.items():
        if data["booked"] and data["not_booked"]:
            debug_log.append(f"\n--- EVENT: {event_key} ---")
            debug_log.append(f"Before deduplication: {len(data['booked'])} booked, {len(data['not_booked'])} not-booked")
            
            # Get email addresses of people who successfully booked
            booked_emails = {record.get("Email", "").lower().strip() for record in data["booked"]}
            debug_log.append(f"Booked emails: {sorted(booked_emails)}")
            
            # Log not-booked entries before filtering
            debug_log.append("Not-booked entries before filtering:")
            for i, record in enumerate(data["not_booked"]):
                email = record.get("Email", "").lower().strip()
                contact = record.get("Contact Name", "Unknown")
                reg_num = record.get("Registration Number", "N/A")
                debug_log.append(f"  {i+1}. {contact} ({email}) - Reg: {reg_num}")
            
            # Filter out not_booked entries where the email is already in booked
            original_not_booked_count = len(data["not_booked"])
            removed_entries = []
            kept_entries = []
            
            for record in data["not_booked"]:
                email = record.get("Email", "").lower().strip()
                if email in booked_emails:
                    removed_entries.append(record)
                else:
                    kept_entries.append(record)
            
            data["not_booked"] = kept_entries
            
            # Log what was removed
            if removed_entries:
                debug_log.append("REMOVED (found in booked list):")
                for record in removed_entries:
                    contact = record.get("Contact Name", "Unknown")
                    email = record.get("Email", "Unknown")
                    reg_num = record.get("Registration Number", "N/A")
                    debug_log.append(f"  - {contact} ({email}) - Reg: {reg_num}")
            
            # Log what was kept
            if kept_entries:
                debug_log.append("KEPT (not found in booked list):")
                for record in kept_entries:
                    contact = record.get("Contact Name", "Unknown")
                    email = record.get("Email", "Unknown")
                    reg_num = record.get("Registration Number", "N/A")
                    debug_log.append(f"  - {contact} ({email}) - Reg: {reg_num}")
            else:
                debug_log.append("KEPT: None (all not-booked entries were duplicates)")
            
            removed_count = original_not_booked_count - len(data["not_booked"])
            debug_log.append(f"After deduplication: {len(data['booked'])} booked, {len(data['not_booked'])} not-booked")
            debug_log.append(f"Removed {removed_count} duplicate(s)")
            
            if removed_count > 0:
                print(f"🧹 Removed {removed_count} duplicate registration(s) from '{event_key}' not-booked list")
    
    # Write debug log to file
    debug_filename = f"deduplication_debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    try:
        with open(debug_filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(debug_log))
        print(f"📝 Debug log written to: {debug_filename}")
    except Exception as e:
        print(f"⚠️ Could not write debug log: {e}")

    return event_data
def is_recent_event(event_text):
    """
    Checks if the event string contains a year within the last 2 calendar years (including current).
    """
    current_year = datetime.now().year
    for year in range(current_year - 1, current_year + 1 + 1):
        if str(year) in event_text:
            return True
    return False

def is_future_event(event_text):
    """
    Checks if the event date is today or in the future by parsing the actual date from the event title.
    Event titles typically start with dates like "Oct 17, 2025:" or "Nov 02, 2025:"
    """
    import re
    from datetime import datetime, date
    
    # Look for date pattern at the start of the event title: "Oct 17, 2025:" or "Nov 02, 2025:"
    date_match = re.match(r'^([A-Za-z]{3})\s+(\d{1,2}),\s+(\d{4}):', event_text.strip())
    
    if date_match:
        try:
            month_str, day_str, year_str = date_match.groups()
            
            # Convert month abbreviation to number
            month_map = {
                'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
            }
            
            month = month_map.get(month_str, None)
            if month is None:
                return False
            
            event_date = date(int(year_str), month, int(day_str))
            today = date.today()
            
            return event_date >= today
        except (ValueError, KeyError):
            pass
    
    # Fallback to year-based filtering if date parsing fails
    current_year = datetime.now().year
    for year in range(current_year, current_year + 6):
        if str(year) in event_text:
            return True
    return False

def is_current_year_event(event_text):
    """
    Checks if the event is in the current year by parsing the actual date when possible.
    """
    import re
    from datetime import datetime, date
    
    # Try to parse the actual date first
    date_match = re.match(r'^([A-Za-z]{3})\s+(\d{1,2}),\s+(\d{4}):', event_text.strip())
    
    if date_match:
        try:
            year_str = date_match.group(3)
            return int(year_str) == datetime.now().year
        except ValueError:
            pass
    
    # Fallback to year string search
    current_year = datetime.now().year
    return str(current_year) in event_text

def is_last_year_event(event_text):
    """
    Checks if the event is from the previous year by parsing the actual date when possible.
    """
    import re
    from datetime import datetime, date
    
    # Try to parse the actual date first
    date_match = re.match(r'^([A-Za-z]{3})\s+(\d{1,2}),\s+(\d{4}):', event_text.strip())
    
    if date_match:
        try:
            year_str = date_match.group(3)
            return int(year_str) == (datetime.now().year - 1)
        except ValueError:
            pass
    
    # Fallback to year string search
    last_year = datetime.now().year - 1
    return str(last_year) in event_text

def write_to_excel(grouped_data, filename="blackpug_registrants.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Keep track of used table names to ensure uniqueness
    used_table_names = set()

    for i, (event, event_records) in enumerate(grouped_data.items()):
        print(f"✏️ Writing sheet: {event[:30]}...")
        sheet_name = event[:31]
        safe_sheet_name = ''.join(c for c in sheet_name if c in string.ascii_letters + string.digits + " _-")[:31]
        ws = wb.create_sheet(title=safe_sheet_name)

        # Sanitized base used for unique Excel table display names (must be
        # alphanumeric, <=15 chars to leave room for suffixes). Computed up
        # front so both the booked and not-booked sections can reference it
        # safely even when an event has only one of the two sections.
        base_name = ''.join(c for c in safe_sheet_name if c.isalnum())[:15] or "Event"
        
        # Add event name as header - but don't let it affect column widths
        ws.cell(row=1, column=1, value=f"Event: {event}")
        ws.column_dimensions['A'].width = 20
        ws.append([])  # Empty row for spacing
        
        current_row = 3
        
        # Handle booked registrations
        booked_records = event_records["booked"]
        if booked_records:
            # Get all unique keys from booked records - ensure they're strings
            all_keys = sorted(set().union(*(r.keys() for r in booked_records)))
            all_keys = [str(k) for k in all_keys]  # Convert to strings
            
            # Reorder columns to put Registration Number after Participants
            if "Participants" in all_keys and "Registration Number" in all_keys:
                all_keys.remove("Registration Number")
                participants_index = all_keys.index("Participants")
                all_keys.insert(participants_index + 1, "Registration Number")
            
            # Write headers for booked section - ensure all are strings
            ws.append(all_keys)
            header_row = current_row
            current_row += 1
            
            # Write booked data
            for record in booked_records:
                row_data = []
                for key in all_keys:
                    value = record.get(key, "")
                    if key == "Participants" and isinstance(value, (int, float)):
                        row_data.append(int(value))
                    elif key == "Registration Number" and isinstance(value, (int, float)):
                        row_data.append(int(value))  # Keep registration numbers as integers
                    else:
                        row_data.append(str(value) if value is not None else "")
                ws.append(row_data)
                current_row += 1
            
            # Add total row if Participants column exists
            if "Participants" in all_keys:
                participants_col_index = all_keys.index("Participants")
                total_participants = sum(record.get("Participants", 0) for record in booked_records if isinstance(record.get("Participants"), (int, float)))
                
                sum_row = [""] * len(all_keys)
                sum_row[0] = "TOTAL:"
                sum_row[participants_col_index] = int(total_participants)
                ws.append(sum_row)
                current_row += 1
            
            # Create table for booked registrations (exclude total row)
            end_col = get_column_letter(len(all_keys))
            table_end_row = current_row - 2 if "Participants" in all_keys else current_row - 1
            table_ref = f"A{header_row}:{end_col}{table_end_row}"
            
            # Generate unique table name
            table_name = f"{base_name}Booked{i+1}"
            counter = 1
            original_table_name = table_name
            while table_name in used_table_names:
                table_name = f"{original_table_name}_{counter}"
                counter += 1
            used_table_names.add(table_name)
            
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Handle not booked registrations
        not_booked_records = event_records["not_booked"]
        if not_booked_records:
            # Add spacing and section title
            current_row += 2
            ws.cell(row=current_row, column=1, value="NOT BOOKED REGISTRATIONS:")
            current_row += 1
            
            # Get all unique keys from not booked records
            not_booked_keys = sorted(set().union(*(r.keys() for r in not_booked_records)))
            not_booked_keys = [str(k) for k in not_booked_keys]  # Convert to strings
            
            # Reorder columns to put Registration Number after Participants
            if "Participants" in not_booked_keys and "Registration Number" in not_booked_keys:
                not_booked_keys.remove("Registration Number")
                participants_index = not_booked_keys.index("Participants")
                not_booked_keys.insert(participants_index + 1, "Registration Number")
            
            # Write headers for not booked section - ensure all are strings
            ws.append(not_booked_keys)
            not_booked_header_row = current_row
            current_row += 1
            
            # Write not booked data
            for record in not_booked_records:
                row_data = []
                for key in not_booked_keys:
                    value = record.get(key, "")
                    if key == "Participants" and isinstance(value, (int, float)):
                        row_data.append(int(value))
                    elif key == "Registration Number" and isinstance(value, (int, float)):
                        row_data.append(int(value))  # Keep registration numbers as integers
                    else:
                        row_data.append(str(value) if value is not None else "")
                ws.append(row_data)
                current_row += 1
            
            # Add total row if Participants column exists
            if "Participants" in not_booked_keys:
                participants_col_index = not_booked_keys.index("Participants")
                total_not_booked = sum(record.get("Participants", 0) for record in not_booked_records if isinstance(record.get("Participants"), (int, float)))
                
                sum_row = [""] * len(not_booked_keys)
                sum_row[0] = "NOT BOOKED TOTAL:"
                sum_row[participants_col_index] = int(total_not_booked)
                ws.append(sum_row)
                current_row += 1
            
            # Create table for not booked registrations (exclude total row)
            not_booked_end_col = get_column_letter(len(not_booked_keys))
            not_booked_table_end_row = current_row - 2 if "Participants" in not_booked_keys else current_row - 1
            not_booked_table_ref = f"A{not_booked_header_row}:{not_booked_end_col}{not_booked_table_end_row}"
            
            # Generate unique table name for not booked
            not_booked_table_name = f"{base_name}NotBooked{i+1}"
            counter = 1
            original_not_booked_table_name = not_booked_table_name
            while not_booked_table_name in used_table_names:
                not_booked_table_name = f"{original_not_booked_table_name}_{counter}"
                counter += 1
            used_table_names.add(not_booked_table_name)
            
            not_booked_table = Table(displayName=not_booked_table_name, ref=not_booked_table_ref)
            not_booked_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            not_booked_table.tableStyleInfo = not_booked_style
            ws.add_table(not_booked_table)
        
        # Auto-fit columns but keep first column at reasonable width
        auto_fit_columns(ws)
        ws.column_dimensions['A'].width = 20
    
    print("💾 Saving Excel workbook...")
    wb.save(filename)
    print(f"✅ Excel export complete: {filename}")
    print(f"📁 File saved to: {os.path.abspath(filename)}")

def main():
    event_url = input("🔗 Enter the Black Pug Event URL: ").strip().split("#")[0]
    if not event_url.startswith("http"):
        print("❌ Invalid URL format. Must start with http or https.")
        return

    # Ask user for filtering preference
    current_year = datetime.now().year
    last_year = current_year - 1
    
    print("\n📅 Choose which events to scrape:")
    print(f"1. Future events - from today forward ({current_year} and beyond)")
    print(f"2. This calendar year's events - any events in {current_year}")
    print(f"3. Last year's events - events from {last_year}")
    
    while True:
        choice = input("\nEnter your choice (1, 2, or 3): ").strip()
        if choice == "1":
            filter_option = "future"
            print("✅ Will scrape future events (from today forward)")
            break
        elif choice == "2":
            filter_option = "current_year"
            print(f"✅ Will scrape this calendar year's events ({current_year})")
            break
        elif choice == "3":
            filter_option = "last_year"
            print(f"✅ Will scrape last year's events ({last_year})")
            break
        else:
            print("❌ Invalid choice. Please enter 1, 2, or 3.")

    driver = build_driver()

    try:
        driver.get(event_url)
        time.sleep(3)
        input("🔒 Log in to Black Pug in the browser. Then press ENTER to continue...")
        wait = WebDriverWait(driver, 15)

        print("🔄 Locating user menu...")
        try:
            user_menus = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "caret")))
        except TimeoutException:
            print("❌ Could not find user menu carets. Trying alternative selector...")
            user_menus = driver.find_elements(By.CSS_SELECTOR, ".dropdown-toggle")

        if len(user_menus) < 2:
            print("❌ Could not find the user dropdown. Are you logged in?")
            print(f"Found {len(user_menus)} menu elements")
            return

        print("📋 Opening user menu...")
        try:
            user_menus[1].click()
        except Exception as e:
            print(f"❌ Error clicking user menu: {e}")
            print("Trying to click the last menu item...")
            user_menus[-1].click()

        print("📋 Navigating to activity history...")
        try:
            view_activity = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(text(), 'View Activity')]")))
        except TimeoutException:
            print("❌ Could not find 'View Activity' link. Trying alternative...")
            view_activity = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(@href, 'activity') or contains(text(), 'Activity')]")))
        
        view_activity.click()

        print("🔄 Waiting for activity data...")
        try:
            # Try the most specific text first
            wait.until(EC.visibility_of_element_located((By.XPATH,
                "//div[contains(@class,'modal')]//div[contains(text(),'Summer Camp & Activities History')]")))
            print("✅ Found modal with 'Summer Camp & Activities History'")
        except TimeoutException:
            try:
                # Try the broader text that appears in the screenshot
                wait.until(EC.visibility_of_element_located((By.XPATH,
                    "//div[contains(@class,'modal')]//div[contains(text(),'Summer Camp & Activities')]")))
                print("✅ Found modal with 'Summer Camp & Activities'")
            except TimeoutException:
                try:
                    # Try looking for any modal with "History" in the text
                    wait.until(EC.visibility_of_element_located((By.XPATH,
                        "//div[contains(@class,'modal')]//div[contains(text(),'History')]")))
                    print("✅ Found modal with 'History' text")
                except TimeoutException:
                    # Try looking for any visible modal
                    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".modal")))
                    print("✅ Found modal dialog")
        
        time.sleep(2)  # Give extra time for content to load

        print("📋 Scraping event registrations...")
        grouped_data = extract_event_rows(driver, filter_option)
        
        # Debug: show what data was collected
        for event_name, event_data in grouped_data.items():
            booked_count = len(event_data["booked"])
            not_booked_count = len(event_data["not_booked"])
            print(f"📊 {event_name}: {booked_count} booked, {not_booked_count} not booked")

        print("📂 Writing data to Excel...")
        write_to_excel(grouped_data)

    except TimeoutException as e:
        print(f"❌ Timeout error: {e}")
        print("💡 This usually means the page took too long to load or elements couldn't be found.")
        print("💡 Try refreshing the page manually and make sure you're logged in.")
    except Exception as e:
        print("❌ Error during scraping:", e)
        import traceback
        print("📋 Full error details:")
        traceback.print_exc()
    finally:
        print("🧹 Closing browser...")
        driver.quit()

if __name__ == "__main__":
    main()
